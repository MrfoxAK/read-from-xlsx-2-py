from xml.etree.ElementTree import iselement
import openpyxl
import numpy as np

wb = openpyxl.load_workbook("t.xlsx")

sh1 = wb['Sheet1']

row = sh1.max_row
col = sh1.max_column

print(row)
print(col)

x1 = []

c=0

for i in range(1,row+1):
     for j in range(1,col+1):
          if sh1.cell(i,j).value is None:
               break
          x1.append(sh1.cell(i,j).value)
          c=c+1

r=int(c/row)


x = np.reshape(x1,(row,r))

print(x)

# print(c)
print("\n2nd matrix is \n")
# print(r)
# print(row)
# print(col)


y1=[]

c1=0
for i in range(1,row+1):
     for j in range(r+1,col+1):
          if sh1.cell(i,j).value is None:
               continue
          y1.append(sh1.cell(i,j).value)
          c1=c1+1

r1=int(c1/row)

y = np.reshape(y1,(row,r1))

print(y)










