import numpy as np
import openpyxl
import matplotlib.pyplot as plt
# from Buttons_Gui import threshold

# threshold()
# print(threshold.thro)

def data():
    with open("save.txt") as f:
        for i in f:
            threshold=float(i)
        
    wb = openpyxl.load_workbook("D:\\xampp\\htdocs\\SERVER\\files\\t.xlsx")

    sh1 = wb['Sheet1']

    row = sh1.max_row
    col = sh1.max_column

    # print(row)
    # print(col)

    x1 = []

    c=0

    for i in range(1,row+1):
        for j in range(1,col+1):
            if sh1.cell(i,j).value is None:
                break
            x1.append(sh1.cell(i,j).value)
            c=c+1

    r=int(c/row)


    XZ_plane = np.reshape(x1,(row,r))
    XZ_list = XZ_plane.tolist()
    # print(XZ_plane) 1
    # print(c)

    # print(r)
    # print(row)
    # print(col)

    y1=[]

    c1=0
    for i in range(1,row+1):
        for j in range(r+1,col+1):
            if sh1.cell(i,j).value is None:
                continue
            y1.append(sh1.cell(i,j).value)
            c1=c1+1

    r1=int(c1/row)

    YZ_plane = np.reshape(y1,(row,r1))

    # print(YZ_plane) 2

    x1=[]
    x=[]
    y1=[]
    y=[]
    z1=[]
    z2=[]
    z=[]
    avg=[]
    val_xz=[]
    x1_contour=np.linspace(1,r,r)
    z1_contour=np.linspace(row,1,row)
    y1_contour=np.linspace(r1,1,r1)
    z2_contour=np.linspace(row,1,row)
    val_yz=[]
    # z_count=0

    # input_data_1=np.array([1,2,3,1,4,2,2,5,6,4,2,7,4,5,3])
    # input_data_2=np.array([1,5,3,6,2,4])

    # XZ_plane=input_data_1.reshape(3,5)
    # YZ_plane=input_data_2.reshape(3,2)

    # print(XZ_plane,"\n",YZ_plane,XZ_plane[2][1])

    # threshold=30

    for i in range(0,row):
        for j in range(0,r):
            # z1_contour.append(int(row-i))
            if((XZ_plane[i][j])>threshold):
                val_xz.append(int (XZ_plane[i][j]))
                x1.append(int(j+1))
                z1.append(int(row-i))
                # print(x1,z1," The Element is ",XZ_plane[i][j])


    for i in range(0,row):
        for j in range(0,r1):
            # val_yz[i].append(YZ_plane[row-i-1][r1-j-1])
            if(YZ_plane[i][j]>threshold):
                # val_yz.append(int (YZ_plane[i][j]))
                y1.append(int(r1-j))
                z2.append(int(row-i))
                # print(y1,z2," The Elements are ",YZ_plane[i][j]) 3

    for i in range(0,len(z1)):
        for j in range(0,len(z2)):
            # print(abs(z1[i]-z2[j]))
            if(abs(z1[i]-z2[j])<2):
                temp=(z1[i]+z2[j])/2
                z.append(temp)
                x.append(x1[i])
                y.append(y1[j])
                # avg.append(float((val_xz[i]+val_yz[j])/2))

    # print(x1_contour,y,z1_contour,avg,'\n',val_yz) 4

    # X,Y=np.meshgrid(x1_contour,z1_contour)
    # # Z,Z=np.meshgrid(val_xz,val_xz)
    # print(X)
    # plt.contourf(X ,Y ,XZ_plane,cmap='jet')
    # plt.savefig('C:/Users/Public/Images/CONTOUR1.jpg')

    # X,Y=np.meshgrid(y1_contour,z2_contour)
    # # Z,Z=np.meshgrid(val_xz,val_xz)
    # print(X)
    # plt.contourf(X ,Y ,YZ_plane,cmap='jet')
    # plt.savefig('CONTOUR2.jpg')

    # chart_studio.tools.set_credentials_file(username='D3V1L1210', api_key='H7G5QLAKDXl49qb5L2Kq')

    # data = XZ_plane
    
    # grid_fig = go.Figure(data =
    #     go.Contour(z = data))



    # url1=(py.plot(grid_fig,filename="Contour1",fileopt='ext',auto_open=False))


    # data = YZ_plane
    
    # grid_fig = go.Figure(data =
    #     go.Contour(z = data))



    # url2=(py.plot(grid_fig,filename="Contour2",fileopt='ext',auto_open=False))


    # fig=go.Figure(data=go.Scatter3d(x=x,y=y,z=z,mode="markers"))
    # url3=py.plot(fig, filename="Scatter_3d" ,file_opt="ext",auto_open=False)
    ax = plt.axes(projection='3d')

    ax.scatter3D(x, y, z)

    plt.show()

    # print(url1,"\n",url2,"\n",url3) 5

    # print(var1,"\n",var2)

    # with open('new.html', 'wb') as f1:
    #     with open ('new.html','r') as f2:

    #         contents = f2.read()
    #         print(contents)
    #         soup = BeautifulSoup(contents, 'html.parser')
    #         target = soup.find_all('contour1')
    #         print(target)
    #         for v in target:
    #             v.replace_with(v.replace('contour1',var1))
    #         html=soup.prettify("utf-8")
    #         print(html,'/n',str(soup))
    #         # f1.write(html)
    #         f2.close()
    #     f1.close()
    
    # first = str(var1)
    # sec = str(var2)

    # index1 = first.find('src')
    # index2 = sec.find('src')
    # # print(index1)


    # i1=index1+5
    # i2=index2+5

    # v1=""
    # while first[i1]!="\0":
    #     if first[i1]=="\"":
    #         break
    #     # print(first[i],end="")
    #     v1 = v1+first[i1]
    #     i1=i1+1

    # v2=""
    # while sec[i2]!="\0":
    #     if sec[i2]=="\"":
    #         break
    #     # print(first[i],end="")
    #     v2 = v2+sec[i2]
    #     i2=i2+1


    # l = []
    # l.append(v1)
    # l.append(v2)

    # print(l)


    # with open ('new.html','r') as f2:
    #     contents = f2.read()
    #     f2.close()
    # soup = BeautifulSoup(contents, 'lxml')
    # target = soup.find_all('iframe')
    # # print(str(target),str(soup),"\ncontents->\n",contents)
    # count=0
    # for b in target:
    #     b['src']=l[count]
    #     count+=1
    # # html=soup.prettify("utf-8")
    # with open ('new.html','w') as f:
    #     f.write(str(soup))

    # print(contents)
        
    # app = Flask('testapp')

    # @app.route('/')
    # def index():
    #     return render_template('new.html', variable=var1)



# data()